import json
import ast


from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from asgiref.sync import sync_to_async
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from channels.layers import get_channel_layer
from django.db.models import Count, Q, Sum, ExpressionWrapper, FloatField
from django.db.models.functions import Coalesce
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from django.core.exceptions import ValidationError
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font





from .forms import FolderForm, FolderUserForm, ReferralForm
from admin.queues import response_queue
from .serializers import TaskSerializer
from .models import Referral, User, Task, UserTask, Groups, Channels, Transaction, Folder, FolderUser, Payout



class AsyncLoginRequiredMixin(LoginRequiredMixin):
    async def dispatch(self, request, *args, **kwargs):
        # Создаем синхронную функцию для проверки аутентификации
        def check_auth():
            return request.user.is_authenticated

        # Обертываем синхронную функцию в sync_to_async
        is_authenticated = await sync_to_async(check_auth)()
        if not is_authenticated:
            return HttpResponseRedirect(reverse('login'))  # Замените 'login' на имя вашего URL для страницы входа
        return await super().dispatch(request, *args, **kwargs)


async def get_channel_members_count(channel_id):
    channel_layer = get_channel_layer()

    # Отправляем запрос боту через WebSocket
    await channel_layer.group_send(
        "notifications",  # Группа, к которой подключен бот
        {
            "type": "get_channel_members_count",  # Метод в Consumer
            "message": json.dumps({
                'action': 'get_channel_members_count',
                'channel_id': channel_id,
            }),
        }
    )
    # Ждем ответа от бота
    response = await response_queue.get()
    return response['members_count']




async def send_websocket_notification(chat_id, message_text):
    channel_layer = get_channel_layer()
    try:
        await channel_layer.group_send(
            "notifications",  # Имя группы WebSocket
            {
                "type": "send_notification",  # Метод в Consumer
                "message": {
                    "chat_id": chat_id,
                    "text": message_text,
                },
            },
        )
        print(f"Сообщение отправлено в WebSocket: chat_id={chat_id}, text={message_text}")
    except Exception as e:
        print(f"Ошибка при отправке сообщения в WebSocket: {e}")





class IndexPageView(AsyncLoginRequiredMixin, View):
    async def get(self, request, *args, **kwargs):
        active_tasks = await Task.objects.filter(status=1).acount()
        archived_tasks = await Task.objects.filter(status=0).acount()
        total_users = await User.objects.acount()
        context = {
            'active_tasks': active_tasks,
            'archived_tasks': archived_tasks,
            'total_users': total_users            
        }
        return await sync_to_async(render)(request, 'admin_panel/index.html', context=context)
    
class CreateTaskView(AsyncLoginRequiredMixin, View):
    async def get(self, request, *args, **kwargs):
        # Получаем список каналов асинхронно
        groups = await sync_to_async(list)(Groups.objects.all())
        channels = await sync_to_async(list)(Channels.objects.all())
        context = {'groups': groups, 'channels': channels}
        return await sync_to_async(render)(request, 'admin_panel/create_task.html', context)
    
    async def post(self, request, *args, **kwargs):
        # Собираем данные из формы
        data = {
            'id': request.POST.get('id') or None,  # Добавляем ID, если оно указано
            'name': request.POST.get('name'),
            'link': request.POST.get('link'),
            'channel': request.POST.get('channel'),
            'groups': request.POST.getlist('groups'),
            'reward': request.POST.get('reward'),
            'end_time': request.POST.get('end_time'),
            'example': request.FILES.get('example'),
            'reminder_start_time': request.POST.get('reminder_start_time'),  # Новое поле
            'reminder_end_time': request.POST.get('reminder_end_time'),  # Новое поле
        }

        # Получаем объект канала
        try:
            channel = await Channels.objects.aget(id=data['channel'])
        except Channels.DoesNotExist:
            return await sync_to_async(render)(
                request,
                'admin_panel/create_task.html',
                {'errors': {'channel': 'канал не найден'}}
            )

        # Получаем количество подписчиков канала
        total_required_subscriptions = 0
        required_subscriptions_list = []
        for group in data['groups']:
            db_group = await Groups.objects.aget(id=group)
            required_subscriptions = await get_channel_members_count(db_group.chat_id)
            required_subscriptions = ast.literal_eval(required_subscriptions)
            total_required_subscriptions += len(required_subscriptions)
            required_subscriptions_list += required_subscriptions
        data['required_subscriptions'] = total_required_subscriptions

        # Валидация данных с помощью сериализатора
        serializer = TaskSerializer(data=data)
        try:
            await sync_to_async(serializer.is_valid)(raise_exception=True)
        except Exception as e:
            # Если есть ошибки, возвращаем форму с ошибками
            gr = await sync_to_async(list)(Groups.objects.all())
            ch = await sync_to_async(list)(Channels.objects.all())
            return await sync_to_async(render)(
                request,
                'admin_panel/create_task.html',
                {'errors': e.detail, 'groups': gr, 'channels': ch}
            )



        # Создаем задание асинхронно
        task_data = serializer.validated_data

        if 'end_time' not in task_data or task_data['end_time'] is None:
            task_data['end_time'] = timezone.now() + timedelta(days=1)

        if 'reminder_start_time' not in task_data or task_data['reminder_start_time'] is None:
            task_data['reminder_start_time'] = timezone.now() + timedelta(days=6)

        if 'reminder_end_time' not in task_data or task_data['reminder_end_time'] is None:
            task_data['reminder_end_time'] = timezone.now() + timedelta(days=14)
        
        if 'reward' not in task_data or task_data['reward'] is None:
            task_data['reward'] = 30

        task = await Task.objects.acreate(
            id=task_data.get('id'),  # Используем указанный ID или None (автогенерация)
            name=task_data['name'],
            link=task_data['link'],
            required_subscriptions=task_data['required_subscriptions'],
            reward=task_data['reward'],
            end_time=task_data['end_time'],
            example=task_data['example'],
            reminder_start_time=task_data.get('reminder_start_time'),  # Новое поле
            reminder_end_time=task_data.get('reminder_end_time'),  # Новое поле
            status='1',  # По умолчанию задание активно
            channel=channel,  # Привязываем задачу к каналу
        )

        # Получаем выбранные группы
        group_ids = data['groups']
        groups = await sync_to_async(list)(Groups.objects.filter(id__in=group_ids))

        # Привязываем группы к задаче
        await sync_to_async(task.groups.set)(groups)

        # Получаем пользователей, которые должны выполнить задание
        user_ids = [int(user) for user in required_subscriptions_list]
        users = await sync_to_async(list)(User.objects.filter(user_id__in=user_ids))

        # Создаем записи UserTask для каждого пользователя
        user_tasks = [
            UserTask(user=user, task=task, status='missed')
            for user in users
        ]

        # Используем bulk_create для создания всех записей за один запрос
        await UserTask.objects.abulk_create(user_tasks)

        # Перенаправляем на страницу списка заданий
        return redirect('panel:task_list')
    


@method_decorator(csrf_exempt, name='dispatch')
class TaskListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Получаем все активные задания с аннотациями для подсчета задач по статусам
        tasks = await sync_to_async(list)(
            Task.objects.filter(status='1').annotate(
                completed_count=Count('usertask', filter=Q(usertask__status='approved')),
                rejected_count=Count('usertask', filter=Q(usertask__status='rejected')),
                pending_count=Count('usertask', filter=Q(usertask__status='pending')),
            )
        )

        

        # Подсчет общего количества активных заданий
        active_tasks_count = await Task.objects.filter(status='1').acount()

        # Подсчет общего количества архивных заданий
        archived_tasks_count = await Task.objects.filter(status='0').acount()

        # Подсчет общего количества заданий
        total_tasks_count = await Task.objects.acount()

        # Подготовка контекста для шаблона
        context = {
            'tasks': tasks,  # Уже отфильтрованы по статусу '1' и содержат счетчики
            'active_tasks_count': active_tasks_count,  # Общее количество активных заданий
            'archived_tasks_count': archived_tasks_count,  # Общее количество архивных заданий
            'total_tasks_count': total_tasks_count,  # Общее количество заданий
        }


        # Рендерим шаблон с контекстом
        return await sync_to_async(render)(request, 'admin_panel/tasks.html', context=context)


class TaskDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, task_id):
        try:
            task = await Task.objects.aget(id=task_id)
        except Task.DoesNotExist:
            raise Http404("Task not found")

        all_tasks = await sync_to_async(list)(
            UserTask.objects.filter(
                Q(status='approved') | Q(status='pending') | Q(status='rejected') | Q(status='missed'),
                task=task,
                task__status='0'
            )
        )

        # Разделяем задачи по статусам
        completed_tasks = [t for t in all_tasks if t.status == 'approved']
        pending_tasks = [t for t in all_tasks if t.status == 'pending']
        rejected_tasks = [t for t in all_tasks if t.status == 'rejected']
        missed_tasks = [t for t in all_tasks if t.status == 'missed']

        all_groups = await sync_to_async(list)(Groups.objects.all())

        context = {
            'task': task,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'rejected_tasks': rejected_tasks,
            'missed_tasks': missed_tasks,
            'all_groups': all_groups,
        }

        return await sync_to_async(render)(request, 'admin_panel/task_detail.html', context)


class EditTaskView(AsyncLoginRequiredMixin, View):
    async def post(self, request, task_id):
        # Получаем задачу
        try:
            task = await Task.objects.aget(id=task_id)
        except Task.DoesNotExist:
            raise Http404()
        

        # Обновляем основные данные задания
        task.name = request.POST.get('name')
        task.link = request.POST.get('link')
        task.required_subscriptions = int(request.POST.get('required_subscriptions'))
        task.reward = float(request.POST.get('reward'))
        task.end_time = request.POST.get('end_time')
        task.reminder_start_time = request.POST.get('reminder_start_time')
        task.reminder_end_time = request.POST.get('reminder_end_time')

        # Сохраняем изменения в задаче
        await task.asave()

        # Обновляем список групп
        group_ids = request.POST.getlist('groups')  # Получаем список ID выбранных групп
        groups = await sync_to_async(list)(Groups.objects.filter(id__in=group_ids))  # Получаем объекты групп
        await sync_to_async(task.groups.set)(groups)  # Обновляем связи ManyToMany

        return redirect('panel:task_detail', task_id=task.id)
    

class CompleteTaskView(AsyncLoginRequiredMixin, View):
    async def get(self, request, task_id):
        try: 
            task = await Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            raise Http404()

        # Меняем статус задания на "архив"
        task.status = '0'
        task.end_time = timezone.now()
        await task.asave()

        return redirect('panel:task_detail', task_id=task.id)


class DeleteTaskView(AsyncLoginRequiredMixin, View):
    async def get(self, request, task_id):
        try: 
            task = await Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            raise Http404()

        # Удаляем задание
        await task.adelete()
        return redirect('panel:task_list')


class CheckTaskView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        context = {}

        # Получаем UserTask с подгрузкой user и task
        user_tasks = await sync_to_async(UserTask.objects.select_related('user', 'task').filter)(status='pending')

        # Получаем все группы
        channels = await sync_to_async(list)(Channels.objects.all())

        # Добавляем данные в контекст
        context['user_tasks'] = user_tasks
        context['channels'] = channels


        # Рендерим шаблон асинхронно
        return await sync_to_async(render)(request, 'admin_panel/check_tasks.html', context=context)
    

class CheckTaskDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, task_id):

        try: 
            task_data = await UserTask.objects.aget(id=task_id)
        except UserTask.DoesNotExist:
            raise Http404()
        
        context = {}

        context['task_data'] = task_data
        
        return await sync_to_async(render)(request, 'admin_panel/review_task.html', context)
    
    async def post(self, request, task_id):
        status = request.POST.get('status')
        feedback = request.POST.get('feedback')

        # Получаем задачу асинхронно
        try: 
            task = await UserTask.objects.select_related('user', 'task').aget(id=task_id)
        except UserTask.DoesNotExist:
            raise Http404()

        # Обновляем статус и обратную связь
        task.status = status
        task.feedback = feedback
        user = task.user
        task_obj = task.task
        await task.asave()
        if status == 'approved':
            reward = task_obj.reward
            # Получаем пользователя асинхронно
            user.balance = user.balance + reward
            await user.asave()
            await Transaction.objects.acreate(
                user=user,
                amount=reward,
                type='completed_task',
                comment=f'Успешное выполнение задания {task_obj.id}'
            )

            # Отправляем уведомление через WebSocket
            message_text = (
                f"Ваше задание '{task_obj.name}' #{task_obj.id} было проверено.\n"
                f"Статус: Одобрено ✅\n"
                f"Вознаграждение: {reward} руб.\n"
            )
        elif status == 'rejected':
            message_text = (
                f"Ваше задание '{task_obj.name}' #{task_obj.id} было проверено.\n"
                f"Статус: Отклонено ❌\n"
            )
        if feedback != "None":
            message_text += f"Обратная связь: {feedback}"
        await send_websocket_notification(user.user_id, message_text)


        return redirect('panel:check_tasks')

    

class UserTaskDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, user_task_id):
        try:
            task_data = await UserTask.objects.aget(id=user_task_id)
        except UserTask.DoesNotExist:
            return Http404()

        context = {
            'task_data': task_data,
        }

        return await sync_to_async(render)(request, 'admin_panel/user_task_detail.html', context)


class ArchivedTasksView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Фильтруем задачи по статусу "архив" (status='0')
        tasks = await sync_to_async(list)(Task.objects.filter(status='0'))

        # Добавляем аннотации для подсчета выполненных, отклоненных и ожидающих задач
        tasks = await sync_to_async(list)(
            Task.objects.filter(status='0').annotate(
                completed_count=Count('usertask', filter=Q(usertask__status='approved')),
                rejected_count=Count('usertask', filter=Q(usertask__status='rejected')),
                pending_count=Count('usertask', filter=Q(usertask__status='pending')),
                missed_count=Count('usertask', filter=Q(usertask__status='missed')),
            )
        )

        context = {
            'tasks': tasks,
        }

        return await sync_to_async(render)(request, 'admin_panel/archived_tasks.html', context)

class UserListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        users = await sync_to_async(list)(User.objects.all())
        return await sync_to_async(render)(request, 'admin_panel/users.html', {'users': users})
    
class DeleteUserView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Получаем список всех пользователей
        users = await sync_to_async(list)(User.objects.all())
        return await sync_to_async(render)(request, 'admin_panel/delete_user.html', {'users': users})

    async def post(self, request):
        # Получаем ID пользователя из формы
        user_id = request.POST.get('user_id')
        
        if user_id:
            # Получаем пользователя по ID
            try:
                user = await User.objects.aget(user_id=user_id)
            except User.DoesNotExist:
                raise Http404()
            
            # Удаляем все связанные данные пользователя
            await UserTask.objects.filter(user=user).adelete()
            await Transaction.objects.filter(user=user).adelete()
            await Referral.objects.filter(referrer=user).adelete()
            await Referral.objects.filter(referred=user).adelete()
            await FolderUser.objects.filter(user=user).adelete()
            
            # Удаляем самого пользователя
            await user.adelete()
            
            # Перенаправляем на страницу списка пользователей
            return redirect('panel:user_list')
        
        # Если user_id не был передан, возвращаем ошибку
        users = await sync_to_async(list)(User.objects.all())
        return await sync_to_async(render)(request, 'admin_panel/delete_user.html', {
            'users': users,
            'error': 'Пользователь не выбран.'
        })



@method_decorator(csrf_exempt, name='dispatch')
class AddTaskView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        return render(request, 'admin_panel/add_task.html')

    async def post(self, request):
        channel_name = request.POST.get('channel_name')
        channel_link = request.POST.get('channel_link')
        reward = request.POST.get('reward')
        await Task.objects.acreate(
            channel_name=channel_name, 
            channel_link=channel_link, 
            reward=reward
        )
        return redirect('task_list')

@method_decorator(csrf_exempt, name='dispatch')
class AddUserView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        return render(request, 'admin_panel/add_user.html')

    async def post(self, request):
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        await User.objects.acreate(
            user_id=user_id, 
            username=username
        )
        return redirect('user_list')
    

@method_decorator(csrf_exempt, name='dispatch')
class UserDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, user_id):
        # Получаем пользователя и его задания
        try:
            user = await User.objects.aget(user_id=user_id)
        except User.DoesNotExist:
            raise Http404()

        user_tasks = await sync_to_async(list)(UserTask.objects.filter(user=user).prefetch_related('task'))

        # Получаем информацию о пригласившем пользователе
        referrer = await Referral.objects.select_related('referrer').filter(referred=user).afirst()
        referrer_username = referrer.referrer if referrer else None

        # Получаем список рефералов пользователя
        referrals = await sync_to_async(list)(Referral.objects.filter(referrer=user).select_related('referred'))

        # Разделяем задания на выполненные и текущие
        completed_tasks = [task for task in user_tasks if task.status == 'approved']
        pending_tasks = [task for task in user_tasks if task.status == 'pending']
        rejected_tasks = [task for task in user_tasks if task.status == 'rejected']
        missed_tasks = [task for task in user_tasks if task.status == 'missed' and task.task.status == '0']
        transactions = await sync_to_async(list)(Transaction.objects.filter(user=user).order_by('-timestamp'))
        if (len(completed_tasks) + len(rejected_tasks) + len(missed_tasks)) != 0:
            completion_rate = ((len(completed_tasks) + len(rejected_tasks)) / (len(completed_tasks) + len(rejected_tasks) + len(missed_tasks))) * 100
        else:
            completion_rate = 0
        all_tasks_without_completed = pending_tasks + rejected_tasks + [task for task in user_tasks if task.status == 'missed']

        # Контекст для шаблона
        context = {
            'user': user,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'rejected_tasks': rejected_tasks,
            'missed_tasks': missed_tasks,
            'transactions': transactions,
            'completion_rate': round(completion_rate, 2),
            'all_tasks_without_completed': all_tasks_without_completed,
            'referrer_username': referrer_username,  # Имя пригласившего пользователя
            'referrals': referrals,
        }
        return await sync_to_async(render)(request, 'admin_panel/user_detail.html', context)
    

class AddFundsView(AsyncLoginRequiredMixin, View):
    async def post(self, request, user_id):
        try:
            user = await User.objects.aget(user_id=user_id)
        except User.DoesNotExist:
            raise Http404()

        amount = float(request.POST.get('amount'))
        comment = request.POST.get('comment')

        # Начисляем деньги
        user.balance += amount
        await user.asave()

        # Логируем транзакцию
        await Transaction.objects.acreate(
            user=user,
            amount=amount,
            type='manual_crediting',
            comment=comment,
        )

        # Добавляем сообщение об успехе

        message_text = (
            f"Вам зачислена награда {amount} руб.\n"
            f"Комментарий к начислению: {comment}"
        )
        await send_websocket_notification(user.user_id, message_text)
        return redirect('panel:user_detail', user_id=user.user_id)
    

class ApproveTaskView(AsyncLoginRequiredMixin, View):
    async def post(self, request, user_id):
        task_id = request.POST.get('task')

        try:
            task = await Task.objects.aget(id=task_id)
            user = await User.objects.aget(user_id=user_id)
        except (Task.DoesNotExist, User.DoesNotExist):
            raise Http404()


        usertask = await UserTask.objects.filter(user=user, task=task).afirst()

        if usertask:
            usertask.status = 'approved'
            usertask.feedback = 'Ручное засчитывание задания'
            await usertask.asave()
        else:
            await UserTask.objects.acreate(
                status='approved',
                feedback='Ручное засчитывание задания',
                user=user,
                task=task
            )

        user.balance += task.reward
        await user.asave()

        # Логируем транзакцию
        await Transaction.objects.acreate(
            user=user,
            amount=task.reward,
            type='completed_task',
            comment=f'Ручное засчитывание задания {task.id}',
        )

        # Отправляем уведомление пользователю
        message_text = (
            f"Ваше задание '{task.name}' #{task.id} было засчитано.\n"
            f"Статус: Одобрено ✅\n"
            f"Вознаграждение: {task.reward} руб.\n"
        )
        await send_websocket_notification(user.user_id, message_text)

        return redirect('panel:user_detail', user_id=user.user_id)



    

class CreateChannelView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        return await sync_to_async(render)(request, 'admin_panel/create_channel.html')

    async def post(self, request):
        name = request.POST.get('name')

        await Channels.objects.acreate(name=name)
        
        return redirect('panel:channel_list')


class ChannelListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        channels = await sync_to_async(list)(Channels.objects.all())
        return await sync_to_async(render)(request, 'admin_panel/channels.html', {'channels': channels})


class ChannelDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, channel_id):
        try: 
            channel = await Channels.objects.aget(id=channel_id)
        except Channels.DoesNotExist:
            raise Http404()

        return await sync_to_async(render)(request, 'admin_panel/channel_detail.html', {'channel': channel})
    
    async def post(self, request, channel_id):
        name = request.POST.get('channel_name')

        try: 
            channel = await Channels.objects.aget(id=channel_id)
        except Channels.DoesNotExist:
            raise Http404()

        channel.name = name
        await channel.asave()
        return redirect('panel:channel_detail', channel_id=channel_id)
    

class DeleteChannelView(AsyncLoginRequiredMixin, View):
    async def post(self, request, channel_id):
        try: 
            channel = await Channels.objects.aget(id=channel_id)
        except Channels.DoesNotExist:
            raise Http404()

        tasks = await UserTask.objects.filter(status='pending', task__channel=channel).acount()
        if tasks == 0:
            await channel.adelete()
            return redirect('panel:channel_list')
        else:
            return redirect('panel:channel_detail', channel_id=channel_id)






class GroupListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        groups = await sync_to_async(list)(Groups.objects.all())
        return await sync_to_async(render)(request, 'admin_panel/groups.html', {'groups': groups})
    

class CreateGroupView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        return await sync_to_async(render)(request, 'admin_panel/create_group.html')

    async def post(self, request):
        name = request.POST.get('name')
        chat_id = request.POST.get('chat_id')

        # Создаем группа
        await Groups.objects.acreate(name=name, chat_id=chat_id)
        return redirect('panel:group_list')

class GroupDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, group_id):
        try:
            group = await Groups.objects.aget(id=group_id)
        except Groups.DoesNotExist:
            raise Http404()
        
        return await sync_to_async(render)(request, 'admin_panel/ group_detail.html', {'group': group})

class EditGroupView(AsyncLoginRequiredMixin, View):
    async def post(self, request, group_id):
        try:
            group = await Groups.objects.aget(id=group_id)
        except Groups.DoesNotExist:
            raise Http404()

        # Обновляем данные канала
        group.name = request.POST.get('name')
        group.chat_id = request.POST.get('chat_id')
        await group.asave()
        return redirect('panel:group_detail', group_id=group.id)

class DeleteGroupView(AsyncLoginRequiredMixin, View):
    async def get(self, request, group_id):
        try:
            group = await Groups.objects.aget(id=group_id)
        except Groups.DoesNotExist:
            raise Http404()
        
        usertask = await UserTask.objects.filter(task__groups=group, status='pending').acount()
        if usertask == 0:
            await group.adelete()
        return redirect('panel:group_list')
    



class UserGroupStatisticsView(AsyncLoginRequiredMixin, View):
    async def get(self, request, *args, **kwargs):
        # Получаем список всех групп
        groups = await sync_to_async(list)(Groups.objects.all())

        # Получаем выбранную группу из параметра запроса
        selected_group_id = request.GET.get('group_id')
        selected_group = None
        if selected_group_id:
            selected_group = await Groups.objects.aget(id=selected_group_id)

        # Получаем список всех пользователей
        users = await sync_to_async(list)(User.objects.all())

        # Собираем статистику для каждого пользователя
        user_stats = []
        for user in users:
            # Фильтруем задачи пользователя по выбранной группе (если группа выбрана)
            if selected_group:
                user_tasks = await sync_to_async(list)(
                    UserTask.objects.filter(user=user, group=selected_group)
                )
            else:
                user_tasks = await sync_to_async(list)(
                    UserTask.objects.filter(user=user)
                )

            # Считаем общее количество задач
            total_tasks = len(user_tasks)

            # Считаем количество пропущенных задач
            total_missed = await UserTask.objects.filter(user=user, status='missed').acount()

            # Считаем процент выполнения задач
            if total_tasks > 0:
                completion_percentage = (total_tasks - total_missed) / total_tasks * 100
            else:
                completion_percentage = 0

            # Считаем количество пропущенных задач подряд
            consecutive_missed = 0
            max_consecutive_missed = 0
            for task in user_tasks:
                if task.status == 'missed':
                    consecutive_missed += 1
                    if consecutive_missed > max_consecutive_missed:
                        max_consecutive_missed = consecutive_missed
                else:
                    consecutive_missed = 0

            # Добавляем статистику пользователя в список
            user_stats.append({
                'user_id': user.user_id,
                'username': user.username,
                'total_missed': total_missed,
                'total_tasks': total_tasks,
                'completion_percentage': completion_percentage,
                'consecutive_missed': max_consecutive_missed,
            })

        # Сортируем пользователей по пропускам всего и пропускам подряд
        sort_by = request.GET.get('sort_by', 'total_missed')
        if sort_by == 'total_missed':
            user_stats.sort(key=lambda x: x['total_missed'], reverse=True)
        elif sort_by == 'consecutive_missed':
            user_stats.sort(key=lambda x: x['consecutive_missed'], reverse=True)

        # Передаем данные в шаблон
        context = {
            'groups': groups,
            'users': user_stats,
            'selected_group': selected_group,
        }

        return await sync_to_async(render)(request, 'admin_panel/user_group_statistics.html', context)



class FolderListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        folders = await sync_to_async(list)(Folder.objects.all())
        
        # Получаем пользователей, которые не находятся ни в одной папке
        users_not_in_any_folder = await sync_to_async(list)(
            User.objects.filter(user_folders__isnull=True)
        )
        
        context = {
            'folders': folders,
            'users_not_in_any_folder': users_not_in_any_folder,  # Передаем список пользователей
        }
        return await sync_to_async(render)(request, 'admin_panel/folders.html', context)
    


class FolderCreateView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        form = FolderForm()
        return await sync_to_async(render)(request, 'admin_panel/create_folder.html', {'form': form})

    async def post(self, request):
        form = FolderForm(request.POST)
        if await sync_to_async(form.is_valid)():
            await sync_to_async(form.save)()
            return redirect('panel:folder_list')
        return await sync_to_async(render)(request, 'admin_panel/create_folder.html', {'form': form})

class FolderDetailView(AsyncLoginRequiredMixin, View):
    async def get(self, request, folder_id):
        try: 
            folder = await Folder.objects.aget(id=folder_id)
        except Folder.DoesNotExist:
            raise Http404()

        users_in_folder = await sync_to_async(list)(folder.folder_users.all())
        
        # Получаем пользователей, которые не находятся ни в одной папке
        users_not_in_any_folder = await sync_to_async(list)(
            User.objects.filter(user_folders__isnull=True)
        )
        
        context = {
            'folder': folder,
            'users_in_folder': users_in_folder,
            'users_not_in_folder': users_not_in_any_folder,  # Только пользователи, не находящиеся ни в одной папке
        }
        return await sync_to_async(render)(request, 'admin_panel/folder_detail.html', context)

    async def post(self, request, folder_id):
        try: 
            folder = await Folder.objects.aget(id=folder_id)
        except Folder.DoesNotExist:
            raise Http404()
        
        user_ids = request.POST.getlist('users')  # Получаем список выбранных пользователей
        
        for user_id in user_ids:
            try:
                user = await User.objects.aget(user_id=user_id)
            except User.DoesNotExist:
                raise Http404()
            
            # Проверяем, что пользователь не находится ни в одной папке
            if not await FolderUser.objects.filter(user=user).aexists():
                await FolderUser.objects.acreate(folder=folder, user=user)
        
        return redirect('panel:folder_detail', folder_id=folder.id)


class FolderRemoveUsersView(AsyncLoginRequiredMixin, View):
    async def post(self, request, folder_id):
        try: 
            folder = await Folder.objects.aget(id=folder_id)
        except Folder.DoesNotExist:
            raise Http404()
        
        user_ids = request.POST.getlist('users')  # Получаем список выбранных пользователей
        for user_id in user_ids:
            try:
                user = await User.objects.aget(user_id=user_id)
            except User.DoesNotExist:
                raise Http404()
            
            await FolderUser.objects.filter(folder=folder, user=user).adelete()
        return redirect('panel:folder_detail', folder_id=folder.id)
    


class FolderDeleteView(AsyncLoginRequiredMixin, View):
    async def get(self, request, folder_id):
        try: 
            folder = await Folder.objects.aget(id=folder_id)
        except Folder.DoesNotExist:
            raise Http404()
        
        return await sync_to_async(render)(request, 'admin_panel/folder_confirm_delete.html', {'folder': folder})

    async def post(self, request, folder_id):
        try: 
            folder = await Folder.objects.aget(id=folder_id)
        except Folder.DoesNotExist:
            raise Http404()
        
        await folder.adelete()
        return redirect('panel:folder_list')
    

class PayoutView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Получаем все транзакции
        transactions = await sync_to_async(list)(Transaction.objects.select_related('payout_id').all())

        # Получаем список всех каналов
        channels = await sync_to_async(list)(Channels.objects.all())

        # Получаем список всех каналов с количеством заданий, которые еще не были выплачены
        channels_with_tasks = await sync_to_async(list)(
            Channels.objects.annotate(
                tasks_list=Count('task')  # Аннотируем количество заданий для каждого канала
            ).prefetch_related('task_set')  # Используем prefetch_related для оптимизации запросов
        )

        # Создаем словарь для хранения задач по каналам
        channels_dict = {
            channel.name: [task for task in channel.task_set.all()]
            for channel in channels_with_tasks
        }

        # Считаем количество принятых скринов для каждого канала
        channels_total = {}
        for channel_name in channels_dict:
            for task in channels_dict[channel_name]:
                for transaction in transactions:
                    if (transaction.comment.endswith(f' {task.id}')) and (transaction.payout_id is None):
                        if channel_name not in channels_total:
                            channels_total[channel_name] = 1
                        else:
                            channels_total[channel_name] += 1

        # Получаем список всех рефералов
        all_ref = await sync_to_async(list)(Referral.objects.select_related('referrer', 'referred').all())

        # Получаем список всех папок
        folders = await sync_to_async(list)(Folder.objects.all())

        # Для каждой папки получаем список пользователей и их балансы
        folder_data = []
        for folder in folders:
            # Получаем пользователей в папке
            folder_users = await sync_to_async(list)(
                User.objects.filter(user_folders__folder=folder).annotate(
                    total_balance=Sum(
                        'users__amount',  # Используем 'users' вместо 'transactions'
                        filter=Q(
                            Q(users__type='completed_task', users__payout_id__isnull=True) |
                            Q(users__type='manual_crediting')  # Учитываем ручные начисления
                        )
                    )
                )
            )

            # Добавляем реферальные бонусы к балансу пользователей
            for ref_user in all_ref:
                referrer = ref_user.referrer
                for f_user in folder_users:
                    if f_user.total_balance is None:
                        f_user.total_balance = 0.0
                    if referrer == f_user:
                        referred_user = ref_user.referred
                        f_user.total_balance += (referred_user.balance * 0.10)
                        f_user.balance += (referred_user.balance * 0.10)

            # Считаем общую сумму выплаты для папки
            total_payout = sum(user.total_balance or 0 for user in folder_users)

            folder_data.append({
                'folder': folder,
                'users': folder_users,
                'total_payout': round(total_payout, 2),  # Добавляем общую сумму выплаты
            })

        # Контекст для шаблона
        context = {
            'channels': channels,  # Передаем каналы вместо групп
            'channels_total': channels_total,  # Статистика по каналам
            'folders': folders,  # Папки остаются без изменений
            'folder_data': folder_data,  # Данные по папкам
        }

        return await sync_to_async(render)(request, 'admin_panel/payouts.html', context)
    
class ExportPayoutsToExcelView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Создаем новый Excel-файл
        wb = Workbook()
        
        # Лист для статистики по каналам
        ws_channels = wb.active
        ws_channels.title = "Статистика по каналам"
        
        # Заголовки для таблицы по каналам
        ws_channels.append(["Название канала", "Принятых скринов"])
        
        # Получаем данные для статистики по каналам
        transactions = await sync_to_async(list)(Transaction.objects.select_related('payout_id').all())
        channels = await sync_to_async(list)(Channels.objects.all())
        channels_with_tasks = await sync_to_async(list)(
            Channels.objects.annotate(
                tasks_list=Count('task')
            ).prefetch_related('task_set')
        )

        channels_dict = {
            channel.name: [task for task in channel.task_set.all()]
            for channel in channels_with_tasks
        }

        channels_total = {}
        for channel_name in channels_dict:
            for task in channels_dict[channel_name]:
                for transaction in transactions:
                    if (transaction.comment.endswith(f' {task.id}')) and (transaction.payout_id is None):
                        if channel_name not in channels_total:
                            channels_total[channel_name] = 1
                        else:
                            channels_total[channel_name] += 1

        # Заполняем данные по каналам
        for channel in channels:
            ws_channels.append([channel.name, channels_total.get(channel.name, 0)])

        # Получаем данные для статистики по папкам
        all_ref = await sync_to_async(list)(Referral.objects.select_related('referrer', 'referred').all())
        folders = await sync_to_async(list)(Folder.objects.all())

        folder_data = []
        for folder in folders:
            folder_users = await sync_to_async(list)(
                User.objects.filter(user_folders__folder=folder).annotate(
                    total_balance=Sum(
                        'users__amount',
                        filter=Q(
                            Q(users__type='completed_task', users__payout_id__isnull=True) |
                            Q(users__type='manual_crediting')
                        )
                    )
                )
            )

            for ref_user in all_ref:
                referrer = ref_user.referrer
                for f_user in folder_users:
                    if f_user.total_balance is None:
                        f_user.total_balance = 0.0
                    if referrer == f_user:
                        referred_user = ref_user.referred
                        f_user.total_balance += (referred_user.balance * 0.10)
                        f_user.balance += (referred_user.balance * 0.10)

            total_payout = sum(user.total_balance or 0 for user in folder_users)

            folder_data.append({
                'folder': folder,
                'users': folder_users,
                'total_payout': round(total_payout, 2),
            })

        # Создаем отдельный лист для каждой папки
        for folder in folder_data:
            ws_folder = wb.create_sheet(title=folder['folder'].name)  # Название листа = название папки
            
            # Заголовки для таблицы по папке
            ws_folder.append(["Пользователь", "Сумма"])
            
            # Заполняем данные по пользователям
            for user in folder['users']:
                ws_folder.append([f"{user.username} (ID: {user.user_id})", user.total_balance or 0])
            
            # Добавляем общую сумму выплат для папки
            ws_folder.append(["Общая сумма выплат", folder['total_payout']])

        # Настройка стилей для заголовков
        for sheet in wb:
            for col in sheet.columns:
                col[0].font = Font(bold=True)

        # Удаляем пустой лист, созданный по умолчанию
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Сохраняем файл в HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=payouts_statistics.xlsx'
        wb.save(response)

        return response
    


class ResetBalancesView(AsyncLoginRequiredMixin, View):
    async def post(self, request):
        # Получаем дату последней выплаты
        last_payout = await sync_to_async(Payout.objects.order_by('-date').first)()
        last_payout_date = last_payout.date if last_payout else None

        # Получаем текущую дату
        current_date = timezone.now()

        # Создаем новую выплату
        payout = await sync_to_async(Payout.objects.create)(is_paid=True)


        all_ref = await sync_to_async(list)(Referral.objects.select_related('referrer', 'referred').all())
        for ref_user in all_ref:
            referrer = ref_user.referrer
            referred = ref_user.referred
            reward = referred.balance * 0.10
            referrer.balance += reward
            await Transaction.objects.acreate(
                user=referrer,
                amount=reward,
                type='referral_bonus',
                comment=f'Реферальный бонус за заработок пользователя @{referred.username} (ID: {referred.user_id})'
            )
            await referrer.asave()

        # Получаем всех пользователей
        users = await sync_to_async(list)(User.objects.all())

        # Для каждого пользователя
        for user in users:
            # Получаем транзакции пользователя за период
            transactions = await sync_to_async(list)(
                Transaction.objects.filter(
                    user=user,
                    payout_id__isnull=True  # Только невыплаченные транзакции
                )
            )

            completed_tasks = sum(1 for t in transactions if t.type == 'completed_task')

            pending_tasks = await UserTask.objects.filter(user=user, status='pending').acount()

            # Считаем сумму наград за выполненные задания
            completed_tasks_reward = sum(
                transaction.task.reward for transaction in transactions
                if transaction.type == 'completed_task' and hasattr(transaction, 'task')
            )

            # Считаем сумму ручных зачислений
            manual_credits = sum(
                transaction.amount for transaction in transactions
                if (transaction.type == 'manual_crediting' or transaction.type == 'referral_bonus') and not transaction.payout_id
            )


            # Формируем сообщение
            message = (
                f"Отчёт за {last_payout_date.strftime('%d.%m.%Y') if last_payout_date else 'начало'} - {current_date.strftime('%d.%m.%Y')}\n"
                f"Количество верных скринов: {completed_tasks}\n"
                f"На рассмотрении: {pending_tasks}\n"
                f"Добавления: {manual_credits:.2f} ₽:\n\n"
            )
            m_index = 1
            for tr in transactions:
                if tr.type == 'manual_crediting' or tr.type == 'referral_bonus':
                    message += f'{m_index}) {tr.amount:.2f}({tr.comment})\n'
                    m_index += 1
            message += f"\nК выплате: {user.balance:.2f} ₽"

            # Связываем транзакции с новой выплатой
            for transaction in transactions:
                transaction.payout_id = payout
                await sync_to_async(transaction.save)()

            # Обнуляем баланс пользователя
            user.balance = 0.0
            await user.asave()

            # Логируем транзакцию для обнуления баланса
            await Transaction.objects.acreate(
                user=user,
                amount=0,
                type='disbersement',  # Тип транзакции: выплата
                comment='Обнуление баланса',
                payout_id=payout
            )

            # Отправляем сообщение пользователю (заглушка, нужно реализовать отправку)
            await send_websocket_notification(user.user_id, message)


        # Перенаправляем обратно на страницу выплат
        return redirect('panel:payout_list')
    

class SendPayoutInformation(AsyncLoginRequiredMixin, View):
    async def post(self, request):
        # Получаем дату последней выплаты
        last_payout = await Payout.objects.order_by('-date').afirst()
        last_payout_date = last_payout.date if last_payout else None

        # Получаем текущую дату
        current_date = timezone.now()


        # Получаем всех пользователей
        users = await sync_to_async(list)(User.objects.all())

        # Для каждого пользователя
        for user in users:
            # Получаем транзакции пользователя за период
            transactions = await sync_to_async(list)(
                Transaction.objects.filter(
                    user=user,
                    payout_id__isnull=True  # Только невыплаченные транзакции
                )
            )

            completed_tasks = sum(1 for t in transactions if t.type == 'completed_task')

            pending_tasks = await UserTask.objects.filter(user=user, status='pending').acount()

            # Считаем сумму наград за выполненные задания
            completed_tasks_reward = sum(
                transaction.task.reward for transaction in transactions
                if transaction.type == 'completed_task' and hasattr(transaction, 'task')
            )

            # Считаем сумму ручных зачислений
            manual_credits = sum(
                transaction.amount for transaction in transactions
                if transaction.type == 'manual_crediting' and not transaction.payout_id
            )
            
            refs = await sync_to_async(list)(Referral.objects.select_related('referred').filter(referrer=user))

            # Общая сумма к выплате

            # Формируем сообщение
            message = (
                f"Отчёт за {last_payout_date.strftime('%d.%m.%Y') if last_payout_date else 'начало'} - {current_date.strftime('%d.%m.%Y')}\n"
                f"Количество верных скринов: {completed_tasks}\n"
                f"На рассмотрении: {pending_tasks}\n"
                f"Добавления: {manual_credits:.2f} ₽:\n\n"
            )
            m_index = 1
            for tr in transactions:
                if tr.type == 'manual_crediting':
                    message += f'{m_index}) {tr.amount:.2f}({tr.comment})\n'
                    m_index += 1
            refs_total = 0
            if refs:
                for ref in refs:
                    ref_user = ref.referred
                    reward = ref_user.balance * 0.10
                    refs_total += reward
                    message += f'{m_index}) {reward:.2f} (Реферальный бонус за заработок пользователя @{ref_user.username} (ID: {ref_user.user_id})'
                    m_index += 1
            message += f"\nК выплате: {round(user.balance + refs_total, 2)} ₽"


            await send_websocket_notification(user.user_id, message)


        # Перенаправляем обратно на страницу выплат
        return redirect('panel:payout_list')
    


class CreateReferralView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        # Получаем список всех пользователей для выбора
        users = await sync_to_async(list)(User.objects.all())
        user_choices = [(user.user_id, f"{user.username} (ID: {user.user_id})") for user in users]

        # Передаем список пользователей в форму
        form = ReferralForm(user_choices=user_choices)
        context = {
            'users': users,
            'form': form,
        }
        return await sync_to_async(render)(request, 'admin_panel/create_referral.html', context)

    async def post(self, request):
        # Получаем список всех пользователей для выбора
        users = await sync_to_async(list)(User.objects.all())
        user_choices = [(user.user_id, f"{user.username} (ID: {user.user_id})") for user in users]

        # Передаем список пользователей в форму
        form = ReferralForm(request.POST, user_choices=user_choices)
        if await sync_to_async(form.is_valid)():
            try:
                referrer_id = form.cleaned_data['referrer']
                referred_id = form.cleaned_data['referred']

                # Получаем объекты пользователей
                referrer = await User.objects.aget(user_id=referrer_id)
                referred = await User.objects.aget(user_id=referred_id)

                # Создаем реферальную связь
                await Referral.objects.acreate(referrer=referrer, referred=referred)

                return redirect('panel:referral_list')  # Перенаправляем на список рефералов
            except ValidationError as e:
                form.add_error(None, e)  # Добавляем ошибку в форму
        else:
            context = {
                'users': users,
                'form': form,
            }
            return await sync_to_async(render)(request, 'admin_panel/create_referral.html', context)
        

class ReferralListView(AsyncLoginRequiredMixin, View):
    async def get(self, request):
        referrals = await sync_to_async(list)(Referral.objects.select_related('referrer', 'referred').all())
        context = {
            'referrals': referrals,
        }
        return await sync_to_async(render)(request, 'admin_panel/referral_list.html', context)
    

class DeleteReferralView(AsyncLoginRequiredMixin, View):
    async def post(self, request, referral_id):
        # Получаем реферальную связь по ID
        referral = await Referral.objects.aget(id=referral_id)
        # Удаляем реферальную связь
        await referral.adelete()
        return redirect('panel:referral_list')